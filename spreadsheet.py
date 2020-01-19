""" A class to open, read and analyze a workbook """

from openpyxl import load_workbook


class Spreadsheet():
    def __init__(self):
        self.workbook = load_workbook("Excel Sample.xlsx")
        self.sheet_names = self.workbook.sheetnames         # sheet_names are only strings
        self.sheet_count = len(self.sheet_names)
        self.sheets = []                                    # this actually holds individual sheet data
        for i in range(self.sheet_count):
            self.sheets.append({"Sheet": self.workbook.get_sheet_by_name(self.sheet_names[i]),
                                "Max Row": self.workbook.get_sheet_by_name(self.sheet_names[i]).max_row,
                                "Max Column": self.workbook.get_sheet_by_name(self.sheet_names[i]).max_column})





